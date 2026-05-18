import os
import sys
import json
import uuid
import threading
import time as _time
import urllib.request
from datetime import datetime
from functools import wraps

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash

sys.path.insert(0, os.path.dirname(__file__))
from database import init_db, get_connection, row_to_dict, rows_to_list
from pdf_processor import extract_bcp_soles, extract_from_excel, extract_from_text, extract_raw_text

_tmp_base = "/tmp" if (os.environ.get("VERCEL") or os.environ.get("RENDER")) else os.path.dirname(__file__)
UPLOAD_FOLDER = os.path.join(_tmp_base, "uploads")
EXPORT_FOLDER = os.path.join(_tmp_base, "exports")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(EXPORT_FOLDER, exist_ok=True)
ALLOWED_PDF = {"pdf"}
ALLOWED_EXCEL = {"xlsx", "xls", "csv"}

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "viva-cont-2026-xK9#mP@qL2")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 31536000
app.config["PREFERRED_URL_SCHEME"] = "https"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = bool(os.environ.get("RENDER") or os.environ.get("VERCEL"))
app.config["PERMANENT_SESSION_LIFETIME"] = 43200  # 12 horas

_ALLOWED_ORIGINS = [
    "https://vivacont.vivaempresasglobal.com",
    "https://vivaos.vivaempresasglobal.com",
]
if not (os.environ.get("RENDER") or os.environ.get("VERCEL")):
    _ALLOWED_ORIGINS += ["http://localhost:5050", "http://127.0.0.1:5050"]
CORS(app, origins=_ALLOWED_ORIGINS, supports_credentials=True)

# Confiar en X-Forwarded-Proto de Cloudflare/Render para que session cookies funcionen en HTTPS
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

init_db()


@app.after_request
def set_cache_headers(response):
    if "text/html" in response.content_type:
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
    return response


def _self_ping():
    """Mantiene el proceso activo en Render free tier (evita cold starts cada 15 min)."""
    _time.sleep(30)  # dar tiempo a gunicorn para arrancar
    url = "https://vivacont.vivaempresasglobal.com/health"
    while True:
        try:
            urllib.request.urlopen(url, timeout=8)
        except Exception:
            pass
        _time.sleep(180)  # cada 3 minutos — bien por debajo del umbral de 15 min

if os.environ.get("RENDER"):
    threading.Thread(target=_self_ping, daemon=True).start()


def allowed_file(filename, allowed):
    if not filename:
        return False
    return "." in filename and filename.rsplit(".", 1)[1].lower() in allowed


# ── Simple in-memory rate limiter (no extra dependency)
import collections
_login_attempts: dict = collections.defaultdict(list)
_LOGIN_LIMIT = 10       # max attempts
_LOGIN_WINDOW = 300     # in 5-minute window


def _is_rate_limited(ip: str) -> bool:
    now = _time.time()
    attempts = _login_attempts[ip]
    # Remove old entries
    _login_attempts[ip] = [t for t in attempts if now - t < _LOGIN_WINDOW]
    if len(_login_attempts[ip]) >= _LOGIN_LIMIT:
        return True
    _login_attempts[ip].append(now)
    return False


# ─────────────────────────────────────────────────────────
# HELPERS MULTI-TENANT
# ─────────────────────────────────────────────────────────

def cid():
    """Retorna el cliente_id del usuario en sesión (aislamiento de datos)."""
    return session.get("cliente_id")


def is_admin_or_above():
    return session.get("user_rol") in ("super_admin", "admin")


def has_privilege(modulo):
    rol = session.get("user_rol")
    if rol in ("super_admin", "admin"):
        return True
    privs = json.loads(session.get("privilegios") or "[]")
    return modulo in privs


# ─────────────────────────────────────────────────────────
# AUTENTICACIÓN
# ─────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return decorated


@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        ip = request.remote_addr or "unknown"
        if _is_rate_limited(ip):
            error = "Demasiados intentos. Espera 5 minutos antes de intentar de nuevo."
            return render_template("login.html", error=error), 429
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        conn = get_connection()
        try:
            user = conn.execute(
                "SELECT * FROM usuarios WHERE (username=? OR email=?) AND activo=1",
                (username, username)
            ).fetchone()
        finally:
            conn.close()
        if user and check_password_hash(user["password_hash"], password):
            session["user_id"]     = user["id"]
            session["user_name"]   = user["nombre"]
            session["user_rol"]    = user["rol"]
            session["user_email"]  = user["email"]
            session["cliente_id"]  = user["cliente_id"]
            session["privilegios"] = user["privilegios"] or "[]"
            conn2 = get_connection()
            try:
                conn2.execute("UPDATE usuarios SET ultimo_acceso=? WHERE id=?",
                              (datetime.now().isoformat(), user["id"]))
                conn2.commit()
            finally:
                conn2.close()
            # Validate next to prevent open redirect: only allow internal paths
            next_url = request.args.get("next", "")
            if next_url and (next_url.startswith("/") and not next_url.startswith("//")):
                safe_next = next_url
            else:
                safe_next = url_for("dashboard")
            return redirect(safe_next)
        error = "Usuario o contraseña incorrectos"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/health")
def health():
    return jsonify({"status": "ok"}), 200


# ─────────────────────────────────────────────────────────
# API: GESTIÓN DE CLIENTES (solo super_admin)
# ─────────────────────────────────────────────────────────

@app.route("/api/admin/clientes", methods=["GET"])
@login_required
def api_get_clientes():
    if session.get("user_rol") != "super_admin":
        return jsonify({"error": "Sin permisos"}), 403
    conn = get_connection()
    rows = rows_to_list(conn.execute("""
        SELECT c.*, COUNT(u.id) as total_usuarios
        FROM clientes c
        LEFT JOIN usuarios u ON u.cliente_id = c.id AND u.rol != 'super_admin'
        GROUP BY c.id ORDER BY c.created_at DESC
    """).fetchall())
    conn.close()
    return jsonify(rows)


@app.route("/api/admin/clientes", methods=["POST"])
@login_required
def api_crear_cliente():
    if session.get("user_rol") != "super_admin":
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json or {}
    razon_social = d.get("razon_social", "").strip()
    ruc          = d.get("ruc", "").strip()
    email        = d.get("email", "").strip().lower()
    telefono     = d.get("telefono", "").strip()
    plan         = d.get("plan", "basic")
    # Datos del admin inicial
    admin_nombre   = d.get("admin_nombre", "").strip()
    admin_email    = d.get("admin_email", "").strip().lower()
    admin_username = d.get("admin_username", "").strip().lower()
    admin_password = d.get("admin_password", "")

    if not all([razon_social, admin_nombre, admin_email, admin_username, admin_password]):
        return jsonify({"error": "razon_social y datos del admin son requeridos"}), 400
    if len(admin_password) < 8:
        return jsonify({"error": "La contraseña del admin debe tener al menos 8 caracteres"}), 400
    if plan not in ("basic", "standard", "premium", "enterprise"):
        plan = "basic"

    _all_privs = json.dumps(["estados_cuenta","analisis_bancario",
                             "estados_resultados","balance_general","facturador"])
    try:
        conn = get_connection()
        c = conn.cursor()
        # Crear cliente
        c.execute("""
            INSERT INTO clientes (razon_social, ruc, email, telefono, plan)
            VALUES (?,?,?,?,?)
        """, (razon_social, ruc, email, telefono, plan))
        cliente_id = c.lastrowid

        # Crear empresa del cliente
        c.execute("""
            INSERT INTO empresa (cliente_id, ruc, razon_social, email, telefono)
            VALUES (?,?,?,?,?)
        """, (cliente_id, ruc, razon_social, email, telefono))

        # Crear admin del cliente
        c.execute("""
            INSERT INTO usuarios (cliente_id, nombre, email, username, password_hash, rol, privilegios)
            VALUES (?,?,?,?,?,?,?)
        """, (cliente_id, admin_nombre, admin_email, admin_username,
              generate_password_hash(admin_password, method='pbkdf2:sha256:50000'),
              'admin', _all_privs))

        conn.commit()
        conn.close()
        return jsonify({"ok": True, "cliente_id": cliente_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/admin/clientes/<int:cid_param>", methods=["PUT"])
@login_required
def api_update_cliente(cid_param):
    if session.get("user_rol") != "super_admin":
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json or {}
    fields = ["razon_social", "ruc", "email", "telefono", "plan", "activo"]
    sets = ", ".join(f"{f}=?" for f in fields if f in d)
    vals = [d[f] for f in fields if f in d]
    if not sets:
        return jsonify({"error": "Nada que actualizar"}), 400
    conn = get_connection()
    conn.execute(f"UPDATE clientes SET {sets} WHERE id=?", vals + [cid_param])
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/clientes/<int:cid_param>/usuarios", methods=["GET"])
@login_required
def api_get_usuarios_cliente(cid_param):
    if session.get("user_rol") != "super_admin":
        return jsonify({"error": "Sin permisos"}), 403
    conn = get_connection()
    rows = rows_to_list(conn.execute("""
        SELECT id, nombre, email, username, rol, privilegios, activo, ultimo_acceso, created_at
        FROM usuarios WHERE cliente_id=? AND rol != 'super_admin' ORDER BY created_at DESC
    """, (cid_param,)).fetchall())
    conn.close()
    return jsonify(rows)


@app.route("/api/admin/clientes/<int:cid_param>/usuarios", methods=["POST"])
@login_required
def api_crear_usuario_cliente(cid_param):
    if session.get("user_rol") != "super_admin":
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json or {}
    nombre   = d.get("nombre", "").strip()
    email    = d.get("email", "").strip().lower()
    username = d.get("username", "").strip().lower()
    password = d.get("password", "")
    rol_nuevo = d.get("rol", "usuario")
    privs    = d.get("privilegios", ["estados_cuenta","analisis_bancario",
                                     "estados_resultados","balance_general","facturador"])
    if not all([nombre, email, username, password]):
        return jsonify({"error": "Todos los campos son requeridos"}), 400
    if len(password) < 8:
        return jsonify({"error": "Contraseña mínimo 8 caracteres"}), 400
    if rol_nuevo not in ("admin", "usuario"):
        rol_nuevo = "usuario"
    conn = get_connection()
    try:
        conn.execute("""
            INSERT INTO usuarios (cliente_id, nombre, email, username, password_hash, rol, privilegios)
            VALUES (?,?,?,?,?,?,?)
        """, (cid_param, nombre, email, username,
              generate_password_hash(password, method='pbkdf2:sha256:50000'),
              rol_nuevo, json.dumps(privs)))
        conn.commit()
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/clientes/<int:cid_param>/usuarios/<int:uid>", methods=["DELETE"])
@login_required
def api_delete_usuario_cliente(cid_param, uid):
    if session.get("user_rol") != "super_admin":
        return jsonify({"error": "Sin permisos"}), 403
    conn = get_connection()
    try:
        conn.execute("DELETE FROM usuarios WHERE id=? AND cliente_id=?", (uid, cid_param))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/clientes/<int:cid_param>", methods=["DELETE"])
@login_required
def api_delete_cliente(cid_param):
    if session.get("user_rol") != "super_admin":
        return jsonify({"error": "Sin permisos"}), 403
    if cid_param == 1:
        return jsonify({"error": "No puedes eliminar el cliente raíz"}), 400
    conn = get_connection()
    conn.execute("UPDATE clientes SET activo=0 WHERE id=?", (cid_param,))
    conn.execute("UPDATE usuarios SET activo=0 WHERE cliente_id=?", (cid_param,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────
# API: GESTIÓN DE USUARIOS (super_admin y admin)
# ─────────────────────────────────────────────────────────

@app.route("/api/usuarios", methods=["GET"])
@login_required
def api_get_usuarios():
    rol = session.get("user_rol")
    conn = get_connection()
    if rol == "super_admin":
        # Ve todos los usuarios de su propio cliente (Viva)
        rows = rows_to_list(conn.execute("""
            SELECT u.id, u.cliente_id, u.nombre, u.email, u.username, u.rol,
                   u.privilegios, u.activo, u.ultimo_acceso, u.created_at,
                   c.razon_social as cliente_nombre
            FROM usuarios u
            LEFT JOIN clientes c ON c.id = u.cliente_id
            WHERE u.cliente_id = ? AND u.rol != 'super_admin'
            ORDER BY u.created_at DESC
        """, (cid(),)).fetchall())
    elif rol == "admin":
        rows = rows_to_list(conn.execute("""
            SELECT id, cliente_id, nombre, email, username, rol,
                   privilegios, activo, ultimo_acceso, created_at
            FROM usuarios
            WHERE cliente_id=? AND rol='usuario'
            ORDER BY created_at DESC
        """, (cid(),)).fetchall())
    else:
        conn.close()
        return jsonify({"error": "Sin permisos"}), 403
    conn.close()
    return jsonify(rows)


@app.route("/api/usuarios", methods=["POST"])
@login_required
def api_crear_usuario():
    rol_sesion = session.get("user_rol")
    if rol_sesion not in ("super_admin", "admin"):
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json or {}
    nombre   = d.get("nombre", "").strip()
    email    = d.get("admin_email" if "admin_email" in d else "email", "").strip().lower()
    if not email:
        email = d.get("email", "").strip().lower()
    username = d.get("username", "").strip().lower()
    password = d.get("password", "")
    rol_nuevo = d.get("rol", "usuario")
    privs    = d.get("privilegios", ["estados_cuenta","analisis_bancario",
                                     "estados_resultados","balance_general","facturador"])

    # admin solo puede crear 'usuario', no otro 'admin'
    if rol_sesion == "admin" and rol_nuevo != "usuario":
        rol_nuevo = "usuario"

    if not all([nombre, email, username, password]):
        return jsonify({"error": "Todos los campos son requeridos"}), 400
    if len(password) < 8:
        return jsonify({"error": "La contraseña debe tener al menos 8 caracteres"}), 400
    if len(nombre) > 120 or len(email) > 120 or len(username) > 60:
        return jsonify({"error": "Campos demasiado largos"}), 400
    try:
        conn = get_connection()
        try:
            conn.execute("""
                INSERT INTO usuarios (cliente_id, nombre, email, username, password_hash, rol, privilegios)
                VALUES (?,?,?,?,?,?,?)
            """, (cid(), nombre, email, username,
                  generate_password_hash(password, method='pbkdf2:sha256:50000'),
                  rol_nuevo, json.dumps(privs)))
            conn.commit()
        finally:
            conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/api/usuarios/<int:uid>", methods=["PUT"])
@login_required
def api_update_usuario(uid):
    rol_sesion = session.get("user_rol")
    if rol_sesion not in ("super_admin", "admin"):
        return jsonify({"error": "Sin permisos"}), 403
    d = request.json or {}

    # Verificar que el usuario pertenece al mismo cliente
    conn = get_connection()
    target = conn.execute("SELECT cliente_id, rol FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not target or target["cliente_id"] != cid():
        conn.close()
        return jsonify({"error": "Sin permisos"}), 403

    privs = json.dumps(d.get("privilegios", [])) if "privilegios" in d else None

    if d.get("password"):
        if privs is not None:
            conn.execute(
                "UPDATE usuarios SET nombre=?,email=?,activo=?,privilegios=?,password_hash=? WHERE id=?",
                (d["nombre"], d["email"], d.get("activo",1), privs,
                 generate_password_hash(d["password"], method='pbkdf2:sha256:50000'), uid))
        else:
            conn.execute(
                "UPDATE usuarios SET nombre=?,email=?,activo=?,password_hash=? WHERE id=?",
                (d["nombre"], d["email"], d.get("activo",1),
                 generate_password_hash(d["password"], method='pbkdf2:sha256:50000'), uid))
    else:
        if privs is not None:
            conn.execute(
                "UPDATE usuarios SET nombre=?,email=?,activo=?,privilegios=? WHERE id=?",
                (d["nombre"], d["email"], d.get("activo",1), privs, uid))
        else:
            conn.execute(
                "UPDATE usuarios SET nombre=?,email=?,activo=? WHERE id=?",
                (d["nombre"], d["email"], d.get("activo",1), uid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/usuarios/<int:uid>", methods=["DELETE"])
@login_required
def api_delete_usuario(uid):
    rol_sesion = session.get("user_rol")
    if rol_sesion not in ("super_admin", "admin"):
        return jsonify({"error": "Sin permisos"}), 403
    if uid == session.get("user_id"):
        return jsonify({"error": "No puedes eliminarte a ti mismo"}), 400
    conn = get_connection()
    target = conn.execute("SELECT cliente_id FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not target or target["cliente_id"] != cid():
        conn.close()
        return jsonify({"error": "Sin permisos"}), 403
    conn.execute("DELETE FROM usuarios WHERE id=?", (uid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────
# PÁGINAS PRINCIPALES
# ─────────────────────────────────────────────────────────

@app.route("/")
@login_required
def dashboard():
    return render_template("index.html")


@app.route("/estados-cuenta")
@login_required
def estados_cuenta():
    return render_template("estados_cuenta.html")


@app.route("/facturador")
@login_required
def facturador():
    return render_template("facturador.html")


@app.route("/configuracion")
@login_required
def configuracion():
    return render_template("configuracion.html")


@app.route("/analisis-bancario")
@login_required
def analisis_bancario():
    return render_template("analisis_bancario.html")


@app.route("/admin/clientes")
@login_required
def admin_clientes():
    if session.get("user_rol") != "super_admin":
        return redirect(url_for("dashboard"))
    return render_template("admin_clientes.html")


@app.route("/usuarios")
@login_required
def admin_usuarios():
    if not is_admin_or_above():
        return redirect(url_for("dashboard"))
    return render_template("admin_usuarios.html")


# ─────────────────────────────────────────────────────────
# API: DASHBOARD KPIs
# ─────────────────────────────────────────────────────────

@app.route("/api/dashboard/kpis")
@login_required
def api_kpis():
    conn = get_connection()
    c = conn.cursor()

    _cid = cid()
    c.execute("SELECT COALESCE(SUM(CASE WHEN importe > 0 THEN importe ELSE 0 END), 0) FROM transacciones WHERE modulo='erp' AND cliente_id=?", (_cid,))
    total_ingresos = c.fetchone()[0]

    c.execute("SELECT COALESCE(SUM(CASE WHEN importe < 0 THEN ABS(importe) ELSE 0 END), 0) FROM transacciones WHERE modulo='erp' AND cliente_id=?", (_cid,))
    total_egresos = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM transacciones WHERE modulo='erp' AND cliente_id=?", (_cid,))
    total_transacciones = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM facturas WHERE estado != 'ANULADA' AND cliente_id=?", (_cid,))
    total_facturas = c.fetchone()[0]

    c.execute("SELECT COALESCE(SUM(total), 0) FROM facturas WHERE estado = 'EMITIDA' AND cliente_id=?", (_cid,))
    total_facturado = c.fetchone()[0]

    c.execute("""
        SELECT mes, SUM(CASE WHEN importe > 0 THEN importe ELSE 0 END) as ingresos,
               SUM(CASE WHEN importe < 0 THEN ABS(importe) ELSE 0 END) as egresos
        FROM transacciones
        WHERE modulo='erp' AND mes != '' AND mes IS NOT NULL AND cliente_id=?
        GROUP BY mes
        ORDER BY MIN(fecha_operacion)
        LIMIT 12
    """, (_cid,))
    flujo_mensual = rows_to_list(c.fetchall())

    c.execute("""
        SELECT tipo, COUNT(*) as cantidad, SUM(ABS(importe)) as monto
        FROM transacciones
        WHERE modulo='erp' AND tipo != '' AND tipo IS NOT NULL AND cliente_id=?
        GROUP BY tipo
    """, (_cid,))
    por_tipo = rows_to_list(c.fetchall())

    c.execute("""
        SELECT * FROM transacciones WHERE modulo='erp' AND cliente_id=?
        ORDER BY created_at DESC LIMIT 10
    """, (_cid,))
    ultimas = rows_to_list(c.fetchall())

    conn.close()
    return jsonify({
        "total_ingresos": round(total_ingresos, 2),
        "total_egresos": round(total_egresos, 2),
        "balance": round(total_ingresos - total_egresos, 2),
        "total_transacciones": total_transacciones,
        "total_facturas": total_facturas,
        "total_facturado": round(total_facturado, 2),
        "flujo_mensual": flujo_mensual,
        "por_tipo": por_tipo,
        "ultimas_transacciones": ultimas,
    })


# ─────────────────────────────────────────────────────────
# API: IMPORTACIÓN (Excel / Drive link)
# ─────────────────────────────────────────────────────────

@app.route("/api/importar/excel", methods=["POST"])
@login_required
def api_importar_excel():
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    file = request.files["file"]
    if not allowed_file(file.filename, ALLOWED_EXCEL):
        return jsonify({"error": "Solo se aceptan archivos .xlsx, .xls, .csv"}), 400

    filename = f"{uuid.uuid4()}_{file.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    result = extract_from_excel(filepath)
    if "error" in result:
        return jsonify({"error": result["error"]}), 500

    _cid = cid()
    conn = get_connection()
    c = conn.cursor()
    inserted = 0
    try:
        for tx in result["transactions"]:
            try:
                c.execute("""
                    INSERT INTO transacciones
                    (cliente_id, modulo, fecha_operacion, referencia, moneda, importe,
                     num_operacion, periodo, banco, fecha, mes, descripcion, tipo, detalle,
                     op, tipo_doc, ruc, cliente_proveedor, num_documento, saldo,
                     doc_cont, comprobante, archivo_origen)
                    VALUES (?,'erp',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    _cid,
                    tx.get("fecha_operacion"), tx.get("referencia"), tx.get("moneda"),
                    tx.get("importe"), tx.get("num_operacion"), tx.get("periodo"),
                    tx.get("banco"), tx.get("fecha"), tx.get("mes"), tx.get("descripcion"),
                    tx.get("tipo"), tx.get("detalle"), tx.get("op"), tx.get("tipo_doc"),
                    tx.get("ruc"), tx.get("cliente_proveedor"), tx.get("num_documento"),
                    tx.get("saldo"), tx.get("doc_cont"), tx.get("comprobante"),
                    tx.get("archivo_origen"),
                ))
                inserted += 1
            except Exception:
                continue

        c.execute("""
            INSERT INTO importaciones (cliente_id, nombre_archivo, tipo_fuente, registros_importados, estado)
            VALUES (?, ?, 'EXCEL', ?, 'COMPLETADO')
        """, (_cid, file.filename, inserted))
        conn.commit()
    finally:
        conn.close()
        try:
            os.remove(filepath)
        except OSError:
            pass

    return jsonify({"success": True, "imported": inserted, "total": result["total"],
                    "nombre_archivo": file.filename})


@app.route("/api/importaciones", methods=["GET"])
@login_required
def api_listar_importaciones():
    conn = get_connection()
    try:
        rows = conn.execute(
            """SELECT id, nombre_archivo, tipo_fuente, registros_importados, estado, created_at
               FROM importaciones WHERE cliente_id=? ORDER BY created_at DESC LIMIT 50""",
            (cid(),)
        ).fetchall()
        return jsonify(rows_to_list(rows))
    finally:
        conn.close()


@app.route("/api/importar/drive", methods=["POST"])
@login_required
def api_importar_drive():
    data = request.get_json() or {}
    url = data.get("url", "").strip()
    if not url:
        return jsonify({"error": "URL requerida"}), 400

    conn = get_connection()
    c = conn.cursor()
    try:
        c.execute("""
            INSERT INTO importaciones (cliente_id, nombre_archivo, tipo_fuente, url_fuente, registros_importados, estado)
            VALUES (?, ?, 'GOOGLE_DRIVE', ?, 0, 'PENDIENTE')
        """, (cid(), url.split("/")[-1][:100], url))
        conn.commit()
        import_id = c.lastrowid
    finally:
        conn.close()

    return jsonify({
        "success": True,
        "message": "Link de Drive registrado. Los datos se procesarán en el módulo de Estados de Cuenta.",
        "import_id": import_id,
        "url": url,
    })


# ─────────────────────────────────────────────────────────
# API: ESTADOS DE CUENTA BANCARIOS
# ─────────────────────────────────────────────────────────

@app.route("/api/estados-cuenta/upload", methods=["POST"])
@login_required
def api_upload_pdf():
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    file = request.files["file"]
    if not allowed_file(file.filename, ALLOWED_PDF | ALLOWED_EXCEL):
        return jsonify({"error": "Solo se aceptan archivos PDF o Excel"}), 400

    banco    = request.form.get("banco", "BCP SOLES")
    filename = f"{uuid.uuid4()}_{file.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    ext = file.filename.rsplit(".", 1)[-1].lower()
    try:
        if ext in ALLOWED_EXCEL:
            result = extract_from_excel(filepath)
        else:
            result = extract_bcp_soles(filepath, banco)

        transactions = result.get("transactions", [])
        # raw_text ya viene embebido en el resultado — sin segunda apertura del PDF
        raw_text = result.get("raw_text", "") or ""

        if not transactions:
            return jsonify({
                "success": False, "transactions": [], "total": 0, "no_data": True,
                "debug": result.get("debug", ""),
                "raw_text": raw_text,
                "message": result.get("error", "No se encontraron transacciones. Usa la opción 'Pegar texto del PDF'."),
            })

        return jsonify({
            "success": True,
            "transactions": transactions,
            "total": len(transactions),
            "strategy": result.get("strategy", ""),
            "preview": True,
        })
    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass


@app.route("/api/estados-cuenta/debug-pdf", methods=["POST"])
@login_required
def api_debug_pdf():
    """Devuelve el texto crudo extraído del PDF para diagnóstico."""
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    file = request.files["file"]
    filename = f"{uuid.uuid4()}_{file.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)
    try:
        raw = extract_raw_text(filepath)
        banco = request.form.get("banco", "BCP SOLES")
        result = extract_bcp_soles(filepath, banco)
    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass
    return jsonify({
        "raw_text": raw,
        "raw_length": len(raw),
        "raw_preview": raw[:2000],
        "transactions_found": result.get("total", 0),
        "strategy": result.get("strategy", "none"),
        "debug": result.get("debug", ""),
    })


@app.route("/api/estados-cuenta/upload-texto", methods=["POST"])
@login_required
def api_upload_texto():
    """Procesa texto copiado del PDF directamente (fallback universal)."""
    data  = request.get_json() or {}
    texto = data.get("texto", "").strip()
    banco = data.get("banco", "BCP SOLES")
    if not texto:
        return jsonify({"error": "Texto vacío"}), 400
    result = extract_from_text(texto, banco)
    transactions = result.get("transactions", [])
    if not transactions:
        return jsonify({
            "success": False,
            "total": 0,
            "message": "No se reconocieron transacciones en el texto. Verifica que contenga fechas (DD/MM/YYYY) y montos.",
        })
    return jsonify({"success": True, "transactions": transactions, "total": len(transactions)})


@app.route("/api/estados-cuenta/confirmar", methods=["POST"])
@login_required
def api_confirmar_transacciones():
    data = request.get_json() or {}
    transactions = data.get("transactions", [])
    periodo_label = (data.get("periodo_label") or "").strip()

    if not transactions:
        return jsonify({"error": "No hay transacciones para guardar"}), 400

    conn = get_connection()
    c = conn.cursor()
    inserted = 0

    # Crear registro de período para que aparezca en las tabs
    meses_list = sorted({tx.get("mes", "") for tx in transactions if tx.get("mes")})
    mes_label  = meses_list[0] if len(meses_list) == 1 else (" - ".join(meses_list[:2]) if meses_list else "")
    anio = 0
    for tx in transactions:
        f = str(tx.get("fecha_operacion") or "")
        if len(f) >= 4 and f[:4].isdigit():
            anio = int(f[:4]); break
    banco_det = transactions[0].get("banco", "") if transactions else ""
    label = periodo_label or f"{mes_label} {anio} · {banco_det}".strip(" ·") or "Importación"
    total_ing = sum(float(tx.get("importe") or 0) for tx in transactions if float(tx.get("importe") or 0) > 0)
    total_egr = abs(sum(float(tx.get("importe") or 0) for tx in transactions if float(tx.get("importe") or 0) < 0))

    c.execute("""
        INSERT INTO periodos_cargados
        (cliente_id, label, mes, anio, banco, archivo, total_transacciones,
         total_ingresos, total_egresos)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, (cid(), label, mes_label, anio, banco_det, label, len(transactions),
          round(total_ing, 2), round(total_egr, 2)))
    periodo_id = c.lastrowid

    for tx in transactions:
        try:
            c.execute("""
                INSERT INTO transacciones
                (cliente_id, modulo, periodo_id, fecha_operacion, referencia, moneda, importe,
                 num_operacion, periodo, banco, fecha, mes, descripcion, tipo, detalle, op,
                 tipo_doc, ruc, cliente_proveedor, num_documento, saldo,
                 doc_cont, comprobante, archivo_origen)
                VALUES (?, 'banco',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                cid(), periodo_id,
                tx.get("fecha_operacion"), tx.get("referencia"), tx.get("moneda"),
                tx.get("importe"), tx.get("num_operacion"), tx.get("periodo"),
                tx.get("banco"), tx.get("fecha"), tx.get("mes"), tx.get("descripcion"),
                tx.get("tipo"), tx.get("detalle"), tx.get("op"), tx.get("tipo_doc"),
                tx.get("ruc"), tx.get("cliente_proveedor"), tx.get("num_documento"),
                tx.get("saldo"), tx.get("doc_cont"), tx.get("comprobante"),
                tx.get("archivo_origen"),
            ))
            inserted += 1
        except Exception:
            continue

    conn.commit()
    conn.close()
    return jsonify({"success": True, "saved": inserted, "periodo_id": periodo_id, "label": label})


@app.route("/api/estados-cuenta/transacciones")
@login_required
def api_get_transacciones():
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 50))
    banco = request.args.get("banco", "")
    mes = request.args.get("mes", "")
    tipo = request.args.get("tipo", "")
    search = request.args.get("search", "")
    offset = (page - 1) * per_page

    conn = get_connection()
    c = conn.cursor()

    periodo_id = request.args.get("periodo_id", "")
    conditions = ["modulo='banco'", "cliente_id=?"]
    params = [cid()]
    if periodo_id and periodo_id != "all":
        conditions.append("periodo_id = ?")
        params.append(int(periodo_id))
    if banco:
        conditions.append("banco = ?")
        params.append(banco)
    if mes:
        conditions.append("mes = ?")
        params.append(mes)
    if tipo:
        conditions.append("tipo = ?")
        params.append(tipo)
    if search:
        conditions.append("(descripcion LIKE ? OR cliente_proveedor LIKE ? OR ruc LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])

    where = f"WHERE {' AND '.join(conditions)}"

    c.execute(f"SELECT COUNT(*) FROM transacciones {where}", params)
    total = c.fetchone()[0]

    c.execute(f"""
        SELECT * FROM transacciones {where}
        ORDER BY fecha_operacion DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset])

    rows = rows_to_list(c.fetchall())
    conn.close()

    return jsonify({"data": rows, "total": total, "page": page, "per_page": per_page})


@app.route("/api/estados-cuenta/transacciones/<int:tx_id>", methods=["PUT"])
@login_required
def api_update_transaccion(tx_id):
    data = request.get_json() or {}
    fields = [
        "tipo_doc", "ruc", "cliente_proveedor", "num_documento",
        "doc_cont", "comprobante", "tipo", "detalle", "descripcion"
    ]
    sets = ", ".join(f"{f} = ?" for f in fields if f in data)
    vals = [data[f] for f in fields if f in data]
    if not sets:
        return jsonify({"error": "Nada que actualizar"}), 400

    conn = get_connection()
    conn.execute(f"UPDATE transacciones SET {sets} WHERE id=? AND cliente_id=?", vals + [tx_id, cid()])
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/estados-cuenta/exportar")
@login_required
def api_exportar_excel():
    import pandas as pd
    import io

    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM transacciones WHERE cliente_id=? ORDER BY fecha_operacion", (cid(),))
    rows = rows_to_list(c.fetchall())
    conn.close()

    if not rows:
        return jsonify({"error": "No hay datos para exportar"}), 404

    df = pd.DataFrame(rows)
    rename = {
        "fecha_operacion": "F.Operac.", "referencia": "Referencia",
        "moneda": "Moneda", "importe": "Importe", "num_operacion": "Num. Ope",
        "periodo": "PERIODO", "banco": "BANCO", "fecha": "FECHA", "mes": "MES",
        "descripcion": "DESCRIPCION", "tipo": "TIPO", "detalle": "DETALLE",
        "op": "OP", "tipo_doc": "TIPO DOC", "ruc": "RUC",
        "cliente_proveedor": "CLIENTES/ PROVEEDOR", "num_documento": "N° DOCUMENTO",
        "saldo": "SALDO", "doc_cont": "DOC CONT", "comprobante": "COMPROBANTE",
    }
    df = df.rename(columns=rename)
    cols = [c for c in rename.values() if c in df.columns]
    df = df[cols]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="EC SOLES 064", index=False)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"VIVA_CONT_Estados_Cuenta_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )


# ─────────────────────────────────────────────────────────
# API: FACTURADOR ELECTRÓNICO
# ─────────────────────────────────────────────────────────

@app.route("/api/facturas", methods=["GET"])
@login_required
def api_get_facturas():
    page = int(request.args.get("page", 1))
    per_page = int(request.args.get("per_page", 20))
    estado = request.args.get("estado", "")
    offset = (page - 1) * per_page

    conn = get_connection()
    c = conn.cursor()

    conditions = ["f.cliente_id=?"]
    params = [cid()]
    if estado:
        conditions.append("f.estado=?")
        params.append(estado)
    where = "WHERE " + " AND ".join(conditions)

    c.execute(f"SELECT COUNT(*) FROM facturas f {where}", params)
    total = c.fetchone()[0]

    c.execute(f"""
        SELECT f.*, GROUP_CONCAT(fi.descripcion, ' | ') as items_desc
        FROM facturas f
        LEFT JOIN factura_items fi ON fi.factura_id = f.id
        {where}
        GROUP BY f.id
        ORDER BY f.created_at DESC
        LIMIT ? OFFSET ?
    """, params + [per_page, offset])

    rows = rows_to_list(c.fetchall())
    conn.close()
    return jsonify({"data": rows, "total": total, "page": page})


@app.route("/api/facturas", methods=["POST"])
@login_required
def api_crear_factura():
    data = request.get_json() or {}
    items = data.pop("items", [])

    conn = get_connection()
    c = conn.cursor()

    subtotal = sum(float(i.get("valor_venta", 0)) for i in items)
    igv = round(subtotal * 0.18, 2)
    total = round(subtotal + igv, 2)

    c.execute("""
        INSERT INTO facturas
        (cliente_id, serie, correlativo, tipo_comprobante, fecha_emision, fecha_vencimiento,
         ruc_emisor, razon_social_emisor, direccion_emisor,
         ruc_cliente, razon_social_cliente, direccion_cliente,
         moneda, subtotal, igv, total, estado, observaciones)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        cid(),
        data.get("serie", "F001"),
        data.get("correlativo", "1"),
        data.get("tipo_comprobante", "FACTURA"),
        data.get("fecha_emision", datetime.now().strftime("%Y-%m-%d")),
        data.get("fecha_vencimiento"),
        data.get("ruc_emisor"), data.get("razon_social_emisor"), data.get("direccion_emisor"),
        data.get("ruc_cliente"), data.get("razon_social_cliente"), data.get("direccion_cliente"),
        data.get("moneda", "PEN"),
        subtotal, igv, total,
        data.get("estado", "BORRADOR"),
        data.get("observaciones"),
    ))
    factura_id = c.lastrowid

    for item in items:
        valor_venta = float(item.get("cantidad", 1)) * float(item.get("precio_unitario", 0))
        igv_item = round(valor_venta * 0.18, 2)
        c.execute("""
            INSERT INTO factura_items
            (factura_id, descripcion, cantidad, unidad, precio_unitario,
             descuento, valor_venta, igv_item, precio_total)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            factura_id,
            item.get("descripcion"), float(item.get("cantidad", 1)),
            item.get("unidad", "UND"), float(item.get("precio_unitario", 0)),
            float(item.get("descuento", 0)), valor_venta, igv_item,
            round(valor_venta + igv_item, 2),
        ))

    conn.commit()
    conn.close()
    return jsonify({"success": True, "id": factura_id, "total": total})


@app.route("/api/facturas/<int:fid>", methods=["GET"])
@login_required
def api_get_factura(fid):
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM facturas WHERE id=? AND cliente_id=?", (fid, cid()))
    factura = row_to_dict(c.fetchone())
    if not factura:
        conn.close()
        return jsonify({"error": "Factura no encontrada"}), 404
    c.execute("SELECT * FROM factura_items WHERE factura_id = ?", (fid,))
    factura["items"] = rows_to_list(c.fetchall())
    conn.close()
    return jsonify(factura)


@app.route("/api/facturas/<int:fid>", methods=["PUT"])
@login_required
def api_actualizar_factura(fid):
    data = request.get_json() or {}
    items = data.pop("items", None)

    conn = get_connection()
    c = conn.cursor()

    fields = [
        "serie", "correlativo", "tipo_comprobante", "fecha_emision", "fecha_vencimiento",
        "ruc_cliente", "razon_social_cliente", "direccion_cliente",
        "moneda", "estado", "observaciones"
    ]
    sets = ", ".join(f"{f} = ?" for f in fields if f in data)
    vals = [data[f] for f in fields if f in data]

    if sets:
        c.execute(f"UPDATE facturas SET {sets} WHERE id=? AND cliente_id=?", vals + [fid, cid()])

    if items is not None:
        c.execute("DELETE FROM factura_items WHERE factura_id = ?", (fid,))
        subtotal = 0
        for item in items:
            valor_venta = float(item.get("cantidad", 1)) * float(item.get("precio_unitario", 0))
            igv_item = round(valor_venta * 0.18, 2)
            subtotal += valor_venta
            c.execute("""
                INSERT INTO factura_items
                (factura_id, descripcion, cantidad, unidad, precio_unitario,
                 descuento, valor_venta, igv_item, precio_total)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (fid, item.get("descripcion"), float(item.get("cantidad", 1)),
                  item.get("unidad", "UND"), float(item.get("precio_unitario", 0)),
                  float(item.get("descuento", 0)), valor_venta, igv_item,
                  round(valor_venta + igv_item, 2)))
        igv = round(subtotal * 0.18, 2)
        c.execute("UPDATE facturas SET subtotal=?, igv=?, total=? WHERE id=?",
                  (round(subtotal, 2), igv, round(subtotal + igv, 2), fid))

    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/facturas/<int:fid>/emitir", methods=["POST"])
@login_required
def api_emitir_factura(fid):
    conn = get_connection()
    conn.execute("UPDATE facturas SET estado='EMITIDA' WHERE id=? AND cliente_id=?", (fid, cid()))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "estado": "EMITIDA"})


@app.route("/api/facturas/<int:fid>/anular", methods=["POST"])
@login_required
def api_anular_factura(fid):
    conn = get_connection()
    conn.execute("UPDATE facturas SET estado='ANULADA' WHERE id=? AND cliente_id=?", (fid, cid()))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "estado": "ANULADA"})


# ─────────────────────────────────────────────────────────
# API: CONFIGURACIÓN EMPRESA
# ─────────────────────────────────────────────────────────

@app.route("/api/empresa", methods=["GET"])
@login_required
def api_get_empresa():
    conn = get_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM empresa WHERE cliente_id=?", (cid(),))
    row = row_to_dict(c.fetchone())
    conn.close()
    return jsonify(row or {})


@app.route("/api/empresa", methods=["PUT"])
@login_required
def api_update_empresa():
    data = request.get_json() or {}
    fields = ["ruc", "razon_social", "nombre_comercial", "direccion", "telefono", "email", "regimen"]
    sets = ", ".join(f"{f} = ?" for f in fields if f in data)
    vals = [data[f] for f in fields if f in data]
    conn = get_connection()
    # Upsert: si no existe empresa para este cliente, la crea
    existing = conn.execute("SELECT id FROM empresa WHERE cliente_id=?", (cid(),)).fetchone()
    if existing:
        conn.execute(f"UPDATE empresa SET {sets}, updated_at=CURRENT_TIMESTAMP WHERE cliente_id=?",
                     vals + [cid()])
    else:
        conn.execute("INSERT INTO empresa (cliente_id) VALUES (?)", (cid(),))
        conn.execute(f"UPDATE empresa SET {sets}, updated_at=CURRENT_TIMESTAMP WHERE cliente_id=?",
                     vals + [cid()])
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ─────────────────────────────────────────────────────────
# API: IMPORTACIÓN DESDE EXCEL DEL TEMPLATE ORIGINAL
# ─────────────────────────────────────────────────────────

@app.route("/api/importar/template-viva", methods=["POST"])
@login_required
def api_importar_template():
    """Importa directamente el template MODELO DE PLANTILLAS VIVA CONT.xlsx"""
    template_path = os.path.join(
        os.path.dirname(__file__), "..",
        "MODELO DE PLANTILLAS VIVA CONT.xlsx"
    )
    if not os.path.exists(template_path):
        return jsonify({"error": "Template no encontrado"}), 404

    result = extract_from_excel(template_path)
    if "error" in result:
        return jsonify({"error": result["error"]}), 500

    conn = get_connection()
    c = conn.cursor()
    inserted = 0
    for tx in result["transactions"]:
        try:
            c.execute("""
                INSERT INTO transacciones
                (cliente_id, modulo, fecha_operacion, referencia, moneda, importe, num_operacion, periodo,
                 banco, fecha, mes, descripcion, tipo, detalle, op, tipo_doc, ruc,
                 cliente_proveedor, num_documento, saldo, doc_cont, comprobante, archivo_origen)
                VALUES (?,'erp',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                cid(),
                tx.get("fecha_operacion"), tx.get("referencia"), tx.get("moneda"),
                tx.get("importe"), tx.get("num_operacion"), tx.get("periodo"),
                tx.get("banco"), tx.get("fecha"), tx.get("mes"), tx.get("descripcion"),
                tx.get("tipo"), tx.get("detalle"), tx.get("op"), tx.get("tipo_doc"),
                tx.get("ruc"), tx.get("cliente_proveedor"), tx.get("num_documento"),
                tx.get("saldo"), tx.get("doc_cont"), tx.get("comprobante"),
                tx.get("archivo_origen"),
            ))
            inserted += 1
        except Exception:
            continue
    conn.commit()
    conn.close()

    return jsonify({"success": True, "imported": inserted})


# ─────────────────────────────────────────────────────────
# API: ANÁLISIS DE ESTADO DE CUENTA (módulo dedicado)
# ─────────────────────────────────────────────────────────

@app.route("/api/analisis-bancario/upload", methods=["POST"])
@login_required
def api_analisis_upload():
    """Sube un PDF bancario, extrae transacciones y devuelve análisis completo."""
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    file = request.files["file"]
    if not allowed_file(file.filename, ALLOWED_PDF | ALLOWED_EXCEL):
        return jsonify({"error": "Solo se aceptan PDF o Excel"}), 400

    banco = request.form.get("banco", "BCP SOLES")
    filename = f"{uuid.uuid4()}_{file.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    ext = file.filename.rsplit(".", 1)[-1].lower()
    orig_filename = file.filename
    try:
        if ext == "pdf":
            result = extract_bcp_soles(filepath, banco)
        else:
            result = extract_from_excel(filepath)

        txs = result.get("transactions", [])
        # raw_text viene embebido — evitamos segunda apertura del PDF
        raw_text = result.get("raw_text", "") or ""
    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass

    if not txs:
        return jsonify({
            "success": False,
            "no_data": True,
            "raw_text": raw_text,
            "debug": result.get("debug", ""),
            "message": result.get("error", "No se encontraron transacciones. Pega el texto del PDF manualmente."),
        })

    analysis = _compute_analysis(txs, banco, orig_filename)
    return jsonify({
        "success": True,
        "transactions": txs,
        "analysis": analysis,
        "filename": orig_filename,
        "raw_text": raw_text,
        "strategy": result.get("strategy", ""),
    })


@app.route("/api/analisis-bancario/upload-texto", methods=["POST"])
@login_required
def api_analisis_texto():
    """Analiza texto copiado directamente del PDF."""
    data  = request.get_json() or {}
    texto = data.get("texto", "").strip()
    banco = data.get("banco", "BCP SOLES")
    if not texto:
        return jsonify({"error": "Texto vacío"}), 400
    result = extract_from_text(texto, banco)
    txs = result.get("transactions", [])
    if not txs:
        return jsonify({
            "success": False,
            "message": "No se reconocieron transacciones. Verifica que el texto contenga fechas (DD/MM/YYYY) y montos.",
        })
    analysis = _compute_analysis(txs, banco, "texto_pegado.txt")
    return jsonify({"success": True, "transactions": txs, "analysis": analysis,
                    "filename": f"Texto pegado · {banco}"})


@app.route("/api/analisis-bancario/from-excel", methods=["POST"])
@login_required
def api_analisis_excel():
    """Analiza el template Excel principal de VIVA CONT."""
    template_path = os.path.join(os.path.dirname(__file__), "..",
                                 "MODELO DE PLANTILLAS VIVA CONT.xlsx")
    if not os.path.exists(template_path):
        return jsonify({"error": "Template no encontrado"}), 404
    result = extract_from_excel(template_path)
    if "error" in result:
        return jsonify({"error": result["error"]}), 500
    txs = result.get("transactions", [])
    analysis = _compute_analysis(txs, "BCP SOLES", "MODELO DE PLANTILLAS VIVA CONT.xlsx")
    return jsonify({"success": True, "transactions": txs, "analysis": analysis,
                    "filename": "MODELO DE PLANTILLAS VIVA CONT.xlsx"})


@app.route("/api/periodos", methods=["GET"])
@login_required
def api_get_periodos():
    """Lista todos los períodos cargados."""
    conn = get_connection()
    rows = rows_to_list(conn.execute(
        "SELECT * FROM periodos_cargados WHERE cliente_id=? ORDER BY anio DESC, created_at DESC",
        (cid(),)
    ).fetchall())
    conn.close()
    return jsonify(rows)


@app.route("/api/periodos/<int:pid>/analysis", methods=["GET"])
@login_required
def api_periodo_analysis(pid):
    """Devuelve el análisis guardado de un período específico."""
    conn = get_connection()
    row = row_to_dict(conn.execute(
        "SELECT * FROM periodos_cargados WHERE id=? AND cliente_id=?", (pid, cid())
    ).fetchone())
    conn.close()
    if not row:
        return jsonify({"error": "Período no encontrado"}), 404
    analysis = json.loads(row["analysis_json"]) if row.get("analysis_json") else {}
    conn2 = get_connection()
    txs = rows_to_list(conn2.execute(
        "SELECT * FROM transacciones WHERE periodo_id=? AND cliente_id=? ORDER BY fecha_operacion",
        (pid, cid())
    ).fetchall())
    conn2.close()
    return jsonify({"periodo": row, "analysis": analysis, "transactions": txs})


@app.route("/api/periodos/consolidado", methods=["GET"])
@login_required
def api_periodo_consolidado():
    """Devuelve análisis consolidado de todos los períodos bancarios."""
    conn = get_connection()
    txs = rows_to_list(conn.execute(
        "SELECT * FROM transacciones WHERE modulo='banco' AND cliente_id=? ORDER BY fecha_operacion",
        (cid(),)
    ).fetchall())
    periodos = rows_to_list(conn.execute(
        "SELECT * FROM periodos_cargados WHERE cliente_id=? ORDER BY anio, created_at",
        (cid(),)
    ).fetchall())
    conn.close()
    if not txs:
        return jsonify({"transactions": [], "analysis": {}, "periodos": []})
    banco = periodos[0]["banco"] if periodos else "Consolidado"
    analysis = _compute_analysis(txs, banco, "Consolidado")
    analysis["banco"] = "Consolidado"
    return jsonify({"transactions": txs, "analysis": analysis, "periodos": periodos})


@app.route("/api/analisis-bancario/guardar", methods=["POST"])
@login_required
def api_analisis_guardar():
    """Persiste las transacciones y crea un período en la BD principal."""
    data = request.get_json() or {}
    txs      = data.get("transactions", [])
    analysis = data.get("analysis", {})
    filename = data.get("filename", "")
    if not txs:
        return jsonify({"error": "Sin transacciones"}), 400

    banco = analysis.get("banco") or (txs[0].get("banco") if txs else "")
    meses_list = sorted({t.get("mes", "") for t in txs if t.get("mes")})
    mes_label  = meses_list[0] if len(meses_list) == 1 else " - ".join(meses_list[:2])
    anio = 0
    for t in txs:
        f = str(t.get("fecha_operacion", "") or "")
        if len(f) >= 4 and f[:4].isdigit():
            anio = int(f[:4]); break
    label = f"{mes_label} {anio} · {banco}".strip(" ·")

    conn = get_connection()
    c = conn.cursor()

    # Crear registro de período
    c.execute("""
        INSERT INTO periodos_cargados
        (cliente_id, label, mes, anio, banco, archivo, total_transacciones,
         total_ingresos, total_egresos, saldo_inicial, saldo_final, analysis_json)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        cid(),
        label, mes_label, anio, banco, filename, len(txs),
        analysis.get("total_ingresos", 0), analysis.get("total_egresos", 0),
        analysis.get("saldo_inicial", 0), analysis.get("saldo_final", 0),
        json.dumps(analysis),
    ))
    periodo_id = c.lastrowid

    inserted = 0
    skipped  = 0
    for tx in txs:
        try:
            c.execute("""
                SELECT COUNT(*) FROM transacciones
                WHERE cliente_id=? AND fecha_operacion=? AND importe=? AND saldo=? AND banco=?
            """, (cid(), tx.get("fecha_operacion"), tx.get("importe"), tx.get("saldo"), tx.get("banco")))
            if c.fetchone()[0] > 0:
                skipped += 1
                continue
            c.execute("""
                INSERT INTO transacciones
                (cliente_id, modulo, periodo_id, fecha_operacion, referencia, moneda, importe,
                 num_operacion, periodo, banco, fecha, mes, descripcion, tipo, detalle,
                 op, tipo_doc, ruc, cliente_proveedor, num_documento, saldo,
                 doc_cont, comprobante, archivo_origen)
                VALUES (?,'banco',?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                cid(),
                periodo_id,
                tx.get("fecha_operacion"), tx.get("referencia"), tx.get("moneda"),
                tx.get("importe"), tx.get("num_operacion"), tx.get("periodo"),
                tx.get("banco"), tx.get("fecha"), tx.get("mes"), tx.get("descripcion"),
                tx.get("tipo"), tx.get("detalle"), tx.get("op"), tx.get("tipo_doc"),
                tx.get("ruc"), tx.get("cliente_proveedor"), tx.get("num_documento"),
                tx.get("saldo"), tx.get("doc_cont"), tx.get("comprobante"),
                tx.get("archivo_origen"),
            ))
            inserted += 1
        except Exception:
            continue
    conn.commit()
    conn.close()
    return jsonify({"success": True, "saved": inserted, "skipped": skipped,
                    "periodo_id": periodo_id, "label": label})


def _safe_float(val, default=0.0):
    """Convierte a float ignorando None, nan y strings inválidos."""
    try:
        v = float(val)
        import math
        return default if math.isnan(v) or math.isinf(v) else v
    except (TypeError, ValueError):
        return default


def _compute_analysis(txs: list, banco: str, filename: str) -> dict:
    """Genera todos los KPIs y series para los gráficos del tablero de análisis."""
    import re
    from collections import defaultdict

    ingresos = sum(_safe_float(t.get("importe")) for t in txs if _safe_float(t.get("importe")) > 0)
    egresos  = abs(sum(_safe_float(t.get("importe")) for t in txs if _safe_float(t.get("importe")) < 0))
    saldos   = [_safe_float(t.get("saldo")) for t in txs if t.get("saldo") not in (None, "", "nan")]
    saldos   = [s for s in saldos if s != 0.0 or True]  # keep zeros too
    saldo_inicial = saldos[0]  if saldos else 0
    saldo_final   = saldos[-1] if saldos else 0

    # ── Flujo diario (saldo acumulado a lo largo del tiempo)
    from datetime import datetime as dt
    dated = []
    for t in txs:
        raw = str(t.get("fecha_operacion") or t.get("fecha") or "")
        if not raw or raw in ("nan", "None", "00:00:00"): continue
        try:
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
                try:
                    d = dt.strptime(raw[:10], fmt)
                    dated.append((d, _safe_float(t.get("importe")), _safe_float(t.get("saldo"))))
                    break
                except ValueError:
                    pass
        except Exception:
            pass
    dated.sort(key=lambda x: x[0])

    flujo_diario = [{"fecha": x[0].strftime("%d/%m"), "importe": round(x[1], 2), "saldo": round(x[2], 2)} for x in dated]

    # ── Por tipo de transacción
    tipo_agg = defaultdict(lambda: {"cantidad": 0, "monto": 0.0})
    for t in txs:
        tipo = t.get("tipo") or "OTRO"
        if str(tipo).strip() in ("nan", "", "None"): tipo = "OTRO"
        tipo_agg[tipo]["cantidad"] += 1
        tipo_agg[tipo]["monto"] += abs(_safe_float(t.get("importe")))
    por_tipo = [{"tipo": k, "cantidad": v["cantidad"], "monto": round(v["monto"], 2)}
                for k, v in sorted(tipo_agg.items(), key=lambda x: -x[1]["monto"])]

    # ── Por mes
    mes_agg = defaultdict(lambda: {"ingresos": 0.0, "egresos": 0.0, "cantidad": 0})
    ORDEN_MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
                   "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    for t in txs:
        mes = t.get("mes") or "—"
        if str(mes).strip() in ("nan", "None", ""): mes = "—"
        imp = _safe_float(t.get("importe"))
        if imp > 0: mes_agg[mes]["ingresos"] += imp
        else:       mes_agg[mes]["egresos"]  += abs(imp)
        mes_agg[mes]["cantidad"] += 1
    por_mes_raw = {k: v for k, v in mes_agg.items()}
    por_mes = [{"mes": m, **por_mes_raw[m]} for m in ORDEN_MESES if m in por_mes_raw]
    if not por_mes:
        por_mes = [{"mes": k, **v} for k, v in sorted(por_mes_raw.items())]
    por_mes = [{"mes": r["mes"], "ingresos": round(r["ingresos"], 2),
                "egresos": round(r["egresos"], 2), "cantidad": r["cantidad"]} for r in por_mes]

    # ── Top 10 proveedores/clientes por monto
    prov_agg = defaultdict(lambda: {"monto": 0.0, "cantidad": 0})
    for t in txs:
        name = (t.get("cliente_proveedor") or "").strip()
        if not name or str(name) in ("nan", "None", ""):
            name = (t.get("descripcion") or "Sin nombre").strip()
        name = name[:40]
        imp = abs(_safe_float(t.get("importe")))
        if imp > 0:
            prov_agg[name]["monto"] += imp
            prov_agg[name]["cantidad"] += 1
    top_proveedores = sorted(
        [{"nombre": k, "monto": round(v["monto"], 2), "cantidad": v["cantidad"]}
         for k, v in prov_agg.items()],
        key=lambda x: -x["monto"]
    )[:10]

    # ── Distribución por tipo de documento
    doc_agg = defaultdict(lambda: {"cantidad": 0, "monto": 0.0})
    for t in txs:
        td = t.get("tipo_doc") or "Sin clasificar"
        if str(td).strip() in ("nan", "", "None"): td = "Sin clasificar"
        doc_agg[td]["cantidad"] += 1
        doc_agg[td]["monto"] += abs(_safe_float(t.get("importe")))
    por_tipo_doc = [{"tipo_doc": k, "cantidad": v["cantidad"], "monto": round(v["monto"], 2)}
                    for k, v in sorted(doc_agg.items(), key=lambda x: -x[1]["monto"])]

    # ── Ingresos vs Egresos acumulados (para gráfico de área)
    acum_ing = 0.0; acum_egr = 0.0
    serie_acum = []
    for x in dated:
        v = _safe_float(x[1])
        if v > 0: acum_ing += v
        else:      acum_egr += abs(v)
        serie_acum.append({"fecha": x[0].strftime("%d/%m"), "ing_acum": round(acum_ing, 2), "egr_acum": round(acum_egr, 2)})

    # ── Categorías de gasto (banco cobros vs terceros)
    cat_banco  = abs(sum(_safe_float(t.get("importe")) for t in txs
                         if "BANCO" in (t.get("cliente_proveedor") or "").upper()
                         or t.get("tipo") == "COBRO"))
    cat_planilla = abs(sum(_safe_float(t.get("importe")) for t in txs
                           if "PLANILLA" in (t.get("detalle") or "").upper()))
    cat_servicios = abs(sum(_safe_float(t.get("importe")) for t in txs
                            if t.get("tipo_doc") in ("RECIBO",)))
    cat_proveedores = abs(sum(_safe_float(t.get("importe")) for t in txs
                              if t.get("tipo_doc") in ("FACTURA", "RHE", "BOLETA")))

    return {
        "banco": banco,
        "archivo": filename,
        "total_transacciones": len(txs),
        "total_ingresos": round(ingresos, 2),
        "total_egresos": round(egresos, 2),
        "balance": round(ingresos - egresos, 2),
        "saldo_inicial": round(saldo_inicial, 2),
        "saldo_final": round(saldo_final, 2),
        "promedio_diario": round((ingresos + egresos) / max(len(set(d[0].strftime("%Y-%m-%d") for d in dated)), 1), 2),
        "flujo_diario": flujo_diario,
        "por_tipo": por_tipo,
        "por_mes": por_mes,
        "top_proveedores": top_proveedores,
        "por_tipo_doc": por_tipo_doc,
        "serie_acumulada": serie_acum,
        "categorias": {
            "Gastos Bancarios": round(cat_banco, 2),
            "Planilla": round(cat_planilla, 2),
            "Servicios": round(cat_servicios, 2),
            "Proveedores": round(cat_proveedores, 2),
        },
    }


# ─────────────────────────────────────────────────────────
# PÁGINAS: ESTADO DE RESULTADOS Y BALANCE GENERAL
# ─────────────────────────────────────────────────────────

@app.route("/estados-resultados")
@login_required
def estados_resultados():
    return render_template("estados_resultados.html")


@app.route("/balance-general")
@login_required
def balance_general():
    return render_template("balance_general.html")


# ─────────────────────────────────────────────────────────
# HELPERS FINANCIEROS
# ─────────────────────────────────────────────────────────

_ER_MAP = {
    "VENTAS NETAS": "ventas_netas",
    "INGRESOS POR VENTAS": "ventas_netas",
    "OTROS INGRESOS": "otros_ingresos",
    "OTROS INGRESOS OPERATIVOS": "otros_ingresos",
    "TOTAL INGRESOS": "total_ingresos",
    "COSTO DE VENTAS": "costo_ventas",
    "COSTO DE SERVICIOS": "costo_ventas",
    "UTILIDAD BRUTA": "utilidad_bruta",
    "GASTOS ADMINISTRATIVOS": "gastos_administrativos",
    "GASTOS DE ADMINISTRACION": "gastos_administrativos",
    "GASTOS DE VENTAS": "gastos_ventas",
    "TOTAL GASTOS OPERATIVOS": "total_gastos_operativos",
    "GASTOS OPERATIVOS": "total_gastos_operativos",
    "EBITDA": "ebitda",
    "DEPRECIACION Y AMORTIZACION": "depreciacion_amortizacion",
    "DEPRECIACION": "depreciacion_amortizacion",
    "EBIT": "ebit",
    "UTILIDAD OPERATIVA": "ebit",
    "GASTOS FINANCIEROS": "gastos_financieros",
    "GASTOS FINANCIEROS NETOS": "gastos_financieros",
    "OTROS GASTOS NETOS": "otros_gastos_netos",
    "OTROS GASTOS": "otros_gastos_netos",
    "UTILIDAD ANTES DE IMPUESTOS": "utilidad_antes_impuestos",
    "UTILIDAD ANTES IMPUESTO": "utilidad_antes_impuestos",
    "IMPUESTO A LA RENTA": "impuesto_renta",
    "IMPUESTO": "impuesto_renta",
    "UTILIDAD NETA": "utilidad_neta",
    "RESULTADO NETO": "utilidad_neta",
}

_BG_MAP = {
    "CAJA Y BANCOS": "caja_bancos",
    "CAJA BANCOS": "caja_bancos",
    "EFECTIVO": "caja_bancos",
    "CUENTAS POR COBRAR": "cuentas_cobrar",
    "CUENTAS COBRAR": "cuentas_cobrar",
    "INVENTARIOS": "inventarios",
    "EXISTENCIAS": "inventarios",
    "OTROS ACTIVOS CORRIENTES": "otros_ac",
    "OTROS ACTIVO CORRIENTE": "otros_ac",
    "TOTAL ACTIVO CORRIENTE": "total_activo_corriente",
    "ACTIVO CORRIENTE": "total_activo_corriente",
    "INMUEBLE MAQUINARIA EQUIPO": "inmueble_maquinaria",
    "INMUEBLE MAQUINARIA Y EQUIPO": "inmueble_maquinaria",
    "INMUEBLES MAQUINARIA Y EQUIPO": "inmueble_maquinaria",
    "PROPIEDAD PLANTA Y EQUIPO": "inmueble_maquinaria",
    "DEPRECIACION ACUMULADA": "depreciacion_acumulada",
    "DEPREC ACUMULADA": "depreciacion_acumulada",
    "ACTIVOS INTANGIBLES": "activos_intangibles",
    "INTANGIBLES": "activos_intangibles",
    "OTROS ACTIVOS NO CORRIENTES": "otros_anc",
    "OTROS ACTIVO NO CORRIENTE": "otros_anc",
    "TOTAL ACTIVO NO CORRIENTE": "total_activo_no_corriente",
    "ACTIVO NO CORRIENTE": "total_activo_no_corriente",
    "TOTAL ACTIVO": "total_activo",
    "ACTIVO TOTAL": "total_activo",
    "CUENTAS POR PAGAR": "cuentas_pagar",
    "CUENTAS PAGAR": "cuentas_pagar",
    "PRESTAMOS CORTO PLAZO": "prestamos_cp",
    "PRESTAMOS CP": "prestamos_cp",
    "OTROS PASIVOS CORRIENTES": "otros_pc",
    "OTROS PASIVO CORRIENTE": "otros_pc",
    "TOTAL PASIVO CORRIENTE": "total_pasivo_corriente",
    "PASIVO CORRIENTE": "total_pasivo_corriente",
    "DEUDA LARGO PLAZO": "deuda_lp",
    "DEUDA LP": "deuda_lp",
    "OTROS PASIVOS NO CORRIENTES": "otros_pnc",
    "OTROS PASIVO NO CORRIENTE": "otros_pnc",
    "TOTAL PASIVO NO CORRIENTE": "total_pasivo_no_corriente",
    "PASIVO NO CORRIENTE": "total_pasivo_no_corriente",
    "TOTAL PASIVO": "total_pasivo",
    "PASIVO TOTAL": "total_pasivo",
    "CAPITAL SOCIAL": "capital_social",
    "CAPITAL": "capital_social",
    "RESERVAS": "reservas",
    "UTILIDADES RETENIDAS": "utilidades_retenidas",
    "RESULTADOS ACUMULADOS": "utilidades_retenidas",
    "RESULTADO DEL EJERCICIO": "resultado_ejercicio",
    "RESULTADO EJERCICIO": "resultado_ejercicio",
    "TOTAL PATRIMONIO": "total_patrimonio",
    "PATRIMONIO": "total_patrimonio",
    "TOTAL PASIVO Y PATRIMONIO": "total_pasivo_patrimonio",
    "TOTAL PASIVO PATRIMONIO": "total_pasivo_patrimonio",
    "TOTAL PASIVO + PATRIMONIO": "total_pasivo_patrimonio",
}


def _parse_financial_excel(filepath, field_map):
    import pandas as pd
    import math
    import re
    import unicodedata

    def clean(v):
        if v is None: return 0.0
        s = str(v).strip().replace(",", "").replace("S/", "").replace("$", "").strip()
        try:
            f = float(s)
            return 0.0 if math.isnan(f) or math.isinf(f) else f
        except (TypeError, ValueError):
            return 0.0

    def norm(s):
        """Normaliza etiqueta: mayúsculas, sin acentos, sin puntuación, sin emojis."""
        s = str(s).upper().strip()
        # Quitar acentos (ó→O, é→E, á→A, etc.)
        s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
        # Quitar paréntesis y su contenido
        s = re.sub(r'\(.*?\)', '', s)
        # Quitar emojis, flechas, guiones especiales
        s = re.sub(r'[▸►→–—✅❌📘⚠️🔵⚫🟣]+', ' ', s)
        # Quitar puntuación (comas, puntos, +, /, etc.)
        s = re.sub(r'[,.:;+/\\!?@#$%^&*\-]', ' ', s)
        # Colapsar espacios
        return re.sub(r'\s+', ' ', s).strip()

    # Pre-normalizar claves del mapa para búsqueda eficiente
    norm_map = {norm(k): v for k, v in field_map.items()}

    result = {"periodo_label": "", "mes": "", "anio": 0, "moneda": "PEN"}
    data = {}

    # Marcadores de sección KPI — al detectarlos se deja de leer campos de datos
    KPI_MARKERS = {"KPI", "INDICADOR", "MARGEN", "TASA", "RATIO", "PUNTO DE EQUILIBRIO",
                   "MULTIPLICADOR", "LIQUIDEZ", "ENDEUDAMIENTO", "SOLVENCIA",
                   "RAZON CORRIENTE", "CAPITAL DE TRABAJO", "DEUDA PATRIMONIO"}

    try:
        df = pd.read_excel(filepath, header=None, dtype=str)
    except Exception as e:
        return {"error": str(e)}

    in_kpi_section = False

    for _, row in df.iterrows():
        cells = [str(c).strip() if (c is not None and str(c).strip() not in ("nan","None","")) else "" for c in row]
        if not any(cells): continue

        col_a = cells[0] if cells else ""
        col_b = cells[1] if len(cells) > 1 else ""
        col_a_clean = norm(col_a)

        # Detectar inicio de sección KPI y dejar de parsear datos financieros
        if any(marker in col_a_clean for marker in KPI_MARKERS):
            in_kpi_section = True

        if col_a_clean == "PERIODO":
            result["periodo_label"] = col_b
        elif col_a_clean == "MES":
            result["mes"] = col_b
        elif col_a_clean in ("ANO", "YEAR"):
            try: result["anio"] = int(float(col_b))
            except: pass
        elif col_a_clean == "MONEDA":
            result["moneda"] = (col_b or "PEN").strip().upper()
        elif not in_kpi_section:
            if col_a_clean in norm_map:
                data[norm_map[col_a_clean]] = clean(col_b)
            else:
                # Coincidencia parcial como fallback (solo fuera de sección KPI)
                for nk, field in norm_map.items():
                    if nk and col_a_clean and (nk in col_a_clean or col_a_clean in nk):
                        data[field] = clean(col_b)
                        break

    result.update(data)
    return result


def _generate_er_template():
    """Genera el Excel template de Estado de Resultados con fórmulas integradas."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "ESTADO DE RESULTADOS"

    # ── Paleta de colores
    BLUE      = PatternFill("solid", start_color="1A3C6E", end_color="1A3C6E")
    PURPLE    = PatternFill("solid", start_color="4C1D95", end_color="4C1D95")
    TEAL      = PatternFill("solid", start_color="0F766E", end_color="0F766E")
    INPUT_BG  = PatternFill("solid", start_color="EFF6FF", end_color="EFF6FF")
    FORM_BG   = PatternFill("solid", start_color="F0FDF4", end_color="F0FDF4")
    TOTAL_BG  = PatternFill("solid", start_color="DBEAFE", end_color="DBEAFE")
    NET_BG    = PatternFill("solid", start_color="1A3C6E", end_color="1A3C6E")
    KPI_BG    = PatternFill("solid", start_color="F5F3FF", end_color="F5F3FF")
    thin      = Side(style="thin", color="CCCCCC")
    border    = Border(bottom=thin)

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 32

    def set_cell(row, col, value, bold=False, color="333333", size=10,
                 fill=None, align="left", fmt=None, italic=False):
        c = ws.cell(row=row, column=col, value=value)
        # Forzar tipo texto si el valor parece fórmula pero es un hint
        if isinstance(value, str) and value.startswith('='):
            c.data_type = 's'
        c.font = Font(bold=bold, color=color, size=size, name="Arial", italic=italic)
        if fill: c.fill = fill
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=(align=="left"))
        if fmt: c.number_format = fmt
        return c

    def section_hdr(row, label, fill=BLUE):
        ws.merge_cells(f'A{row}:D{row}')
        c = ws.cell(row=row, column=1, value=label)
        c.fill = fill
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def input_row(row, label, hint=""):
        set_cell(row, 1, f"  {label}", fill=INPUT_BG)
        c = set_cell(row, 2, 0, color="0000FF", fill=INPUT_BG, align="right", fmt='#,##0.00')
        set_cell(row, 3, "", fill=INPUT_BG)
        set_cell(row, 4, hint, color="888888", size=9, italic=True, fill=INPUT_BG)
        ws.row_dimensions[row].height = 18
        return c

    def formula_row(row, label, formula, hint="", is_net=False, is_subtotal=False):
        fill = NET_BG if is_net else (TOTAL_BG if is_subtotal else FORM_BG)
        txt_color = "FFFFFF" if is_net else ("1A3C6E" if is_subtotal else "065F46")
        set_cell(row, 1, label, bold=True, color=txt_color, fill=fill)
        c = ws.cell(row=row, column=2, value=formula)
        c.font = Font(bold=True, color=txt_color, size=10, name="Arial")
        c.fill = fill
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '#,##0.00'
        set_cell(row, 3, "", fill=fill)
        set_cell(row, 4, hint, color="888888", size=9, italic=True, fill=fill)
        ws.row_dimensions[row].height = 20
        return c

    def kpi_row(row, label, formula, fmt=None, hint=""):
        set_cell(row, 1, f"  {label}", fill=KPI_BG)
        c = ws.cell(row=row, column=2, value=formula)
        c.font = Font(bold=True, color="4C1D95", size=10, name="Arial")
        c.fill = KPI_BG
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = fmt or '0.00%'
        set_cell(row, 3, "", fill=KPI_BG)
        set_cell(row, 4, hint, color="888888", size=9, italic=True, fill=KPI_BG)
        ws.row_dimensions[row].height = 18

    # ── ROW 1: TÍTULO
    ws.merge_cells('A1:D1')
    t = ws.cell(row=1, column=1, value="ESTADO DE RESULTADOS  |  VIVA CONSULTING")
    t.fill = BLUE
    t.font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ── ROWS 2–5: METADATA
    meta = [("PERIODO","","Ej: Enero 2026"),("MES","","Ej: Enero"),
            ("AÑO","","Ej: 2026"),("MONEDA","PEN","PEN · USD")]
    for i, (lbl, val, hint) in enumerate(meta, start=2):
        set_cell(i, 1, lbl, bold=True, color="1A3C6E")
        c = ws.cell(row=i, column=2, value=val)
        c.font = Font(bold=False, color="0000FF", size=10, name="Arial")
        c.alignment = Alignment(horizontal="left", vertical="center")
        set_cell(i, 4, hint, color="AAAAAA", size=9, italic=True)
        ws.row_dimensions[i].height = 18

    # ── ROW 6: SEPARADOR
    ws.row_dimensions[6].height = 6

    # ── ROW 7: CABECERAS
    for col, lbl in [(1,"CONCEPTO"),(2,"MONTO (S/)"),(3,""),(4,"NOTAS / REFERENCIA")]:
        c = ws.cell(row=7, column=col, value=lbl)
        c.fill = BLUE
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.alignment = Alignment(horizontal="center" if col==2 else "left", vertical="center")
    ws.row_dimensions[7].height = 22

    # ── FILAS DE DATOS (filas fijas para fórmulas)
    # Ingresos
    section_hdr(8,  "▸  INGRESOS")
    input_row  (9,  "VENTAS NETAS",                  "Ingresos por ventas de bienes o servicios")
    input_row  (10, "OTROS INGRESOS OPERATIVOS",      "Intereses ganados, devoluciones, etc.")
    formula_row(11, "TOTAL INGRESOS",                 "=B9+B10",              "Ventas + Otros ingresos", is_subtotal=True)

    # Costos
    section_hdr(12, "▸  COSTO DE VENTAS")
    input_row  (13, "COSTO DE VENTAS",                "Materia prima, mano obra directa, etc.")
    formula_row(14, "UTILIDAD BRUTA",                 "=B11-B13",             "Ingresos – Costo Ventas", is_subtotal=True)

    # Gastos Operativos
    section_hdr(15, "▸  GASTOS OPERATIVOS")
    input_row  (16, "GASTOS ADMINISTRATIVOS",         "Sueldos admin, alquileres, servicios")
    input_row  (17, "GASTOS DE VENTAS",               "Comisiones, publicidad, marketing")
    formula_row(18, "TOTAL GASTOS OPERATIVOS",        "=B16+B17",             "G.Admin + G.Ventas", is_subtotal=True)
    formula_row(19, "EBITDA",                         "=B14-B18",             "Utilidad Bruta – Gastos Operativos", is_subtotal=True)
    input_row  (20, "DEPRECIACIÓN Y AMORTIZACIÓN",    "Activos fijos e intangibles del período")
    formula_row(21, "EBIT  (Utilidad Operativa)",     "=B19-B20",             "EBITDA – Depreciación")

    # Otros
    section_hdr(22, "▸  OTROS GASTOS / INGRESOS")
    input_row  (23, "GASTOS FINANCIEROS",             "Intereses de préstamos bancarios")
    input_row  (24, "OTROS GASTOS NETOS",             "Gastos o ingresos extraordinarios")
    formula_row(25, "UTILIDAD ANTES DE IMPUESTOS",    "=B21-B23-B24",         "EBIT – G.Fin – Otros", is_subtotal=True)
    formula_row(26, "IMPUESTO A LA RENTA (29.5%)",   "=IF(B25>0,B25*0.295,0)", "UAI × 29.5% Régimen General")
    formula_row(27, "UTILIDAD NETA",                  "=B25-B26",             "UAI – Impuesto", is_net=True)

    # ── ROW 28: SEPARADOR
    ws.row_dimensions[28].height = 10

    # ── KPIs FINANCIEROS AUTOMÁTICOS
    section_hdr(29, "▸  KPIs FINANCIEROS  (calculados automáticamente)", fill=PURPLE)
    for col, lbl in [(1,"INDICADOR"),(2,"VALOR"),(4,"REFERENCIA / BENCHMARK")]:
        c = ws.cell(row=30, column=col, value=lbl)
        c.fill = KPI_BG
        c.font = Font(bold=True, color="4C1D95", size=9, name="Arial")
        c.alignment = Alignment(horizontal="center" if col==2 else "left", vertical="center")
    ws.row_dimensions[30].height = 18

    kpi_row(31, "Margen Bruto",            "=IF(B11>0,B14/B11,0)",   "0.0%", "Utilidad Bruta / Ingresos  ·  Objetivo: >40%")
    kpi_row(32, "Margen EBITDA",           "=IF(B11>0,B19/B11,0)",   "0.0%", "EBITDA / Ingresos  ·  Objetivo: >15%")
    kpi_row(33, "Margen Operativo (EBIT)", "=IF(B11>0,B21/B11,0)",   "0.0%", "EBIT / Ingresos  ·  Objetivo: >10%")
    kpi_row(34, "Margen Neto",             "=IF(B11>0,B27/B11,0)",   "0.0%", "Utilidad Neta / Ingresos  ·  Objetivo: >8%")
    kpi_row(35, "Costo Ventas / Ingresos", "=IF(B11>0,B13/B11,0)",   "0.0%", "Eficiencia de producción")
    kpi_row(36, "Gastos Op. / Ingresos",   "=IF(B11>0,B18/B11,0)",   "0.0%", "Eficiencia operativa  ·  Objetivo: <40%")
    kpi_row(37, "Tasa Efectiva Impuesto",  "=IF(B25>0,B26/B25,0)",   "0.0%", "Impuesto / UAI  ·  Régimen General: 29.5%")
    kpi_row(38, "Multiplicador Ingresos",  "=IF(B9>0,B11/B9,0)",     "0.00", "Total Ingresos / Ventas Netas")

    # Separador KPIs
    ws.row_dimensions[39].height = 8

    # KPIs monetarios
    section_hdr(40, "▸  INDICADORES ABSOLUTOS", fill=TEAL)
    for row, lbl, frm, hint in [
        (41, "Utilidad Bruta (S/)",         "=B14", "Cobertura sobre costos directos"),
        (42, "EBITDA (S/)",                 "=B19", "Generación operativa de caja"),
        (43, "EBIT (S/)",                   "=B21", "Beneficio antes de intereses e impuestos"),
        (44, "Utilidad Neta (S/)",          "=B27", "Beneficio final del período"),
        (45, "Gastos Totales (S/)",         "=B13+B18+B20+B23+B24+B26", "Suma de todos los gastos"),
        (46, "Punto de Equilibrio (est.)",  "=IF((B11-B13)>0,B18/(1-B13/B11),0)", "Ingresos necesarios para cubrir gastos fijos"),
    ]:
        set_cell(row, 1, f"  {lbl}", fill=KPI_BG)
        c = ws.cell(row=row, column=2, value=frm)
        c.font = Font(bold=True, color="065F46", size=10, name="Arial")
        c.fill = KPI_BG
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '#,##0.00'
        set_cell(row, 4, hint, color="888888", size=9, italic=True, fill=KPI_BG)
        ws.row_dimensions[row].height = 18

    # ── LEYENDA
    ws.row_dimensions[47].height = 10
    ws.merge_cells('A48:D48')
    leg = ws.cell(row=48, column=1,
        value="LEYENDA:  🔵 Celdas AZULES = ingresa tus datos   ·   ⚫ Celdas NEGRAS = calculadas automáticamente por fórmula   ·   🟣 Sección KPIs = solo lectura")
    leg.font = Font(name="Arial", size=9, italic=True, color="666666")
    leg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[48].height = 22

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _generate_bg_template():
    """Genera el Excel template de Balance General con fórmulas y cuadre automático."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "BALANCE GENERAL"

    BLUE     = PatternFill("solid", start_color="1A3C6E", end_color="1A3C6E")
    PURPLE   = PatternFill("solid", start_color="4C1D95", end_color="4C1D95")
    GREEN    = PatternFill("solid", start_color="065F46", end_color="065F46")
    RED      = PatternFill("solid", start_color="7F1D1D", end_color="7F1D1D")
    INPUT_BG = PatternFill("solid", start_color="EFF6FF", end_color="EFF6FF")
    FORM_BG  = PatternFill("solid", start_color="DBEAFE", end_color="DBEAFE")
    TOT_BG   = PatternFill("solid", start_color="BFDBFE", end_color="BFDBFE")
    KPI_BG   = PatternFill("solid", start_color="F5F3FF", end_color="F5F3FF")
    CHK_OK   = PatternFill("solid", start_color="D1FAE5", end_color="D1FAE5")

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 34

    def set_cell(row, col, value, bold=False, color="333333", size=10,
                 fill=None, align="left", fmt=None, italic=False):
        c = ws.cell(row=row, column=col, value=value)
        if isinstance(value, str) and value.startswith('='):
            c.data_type = 's'
        c.font = Font(bold=bold, color=color, size=size, name="Arial", italic=italic)
        if fill: c.fill = fill
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align=="left"))
        if fmt: c.number_format = fmt
        return c

    def sec_hdr(row, label, fill=BLUE):
        ws.merge_cells(f'A{row}:D{row}')
        c = ws.cell(row=row, column=1, value=label)
        c.fill = fill
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 20

    def inp(row, label, hint="", negative=False):
        set_cell(row, 1, f"  {label}", fill=INPUT_BG)
        c = ws.cell(row=row, column=2, value=0)
        c.font = Font(bold=False, color="0000FF", size=10, name="Arial")
        c.fill = INPUT_BG
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '#,##0.00'
        set_cell(row, 3, "", fill=INPUT_BG)
        note = f"{'⚠ Ingresar como valor NEGATIVO' if negative else ''}{'  |  ' if negative and hint else ''}{hint}"
        set_cell(row, 4, note, color="888888", size=9, italic=True, fill=INPUT_BG)
        ws.row_dimensions[row].height = 18

    def frm(row, label, formula, hint="", is_main=False):
        fill = TOT_BG if is_main else FORM_BG
        color = "1A3C6E" if is_main else "065F46"
        set_cell(row, 1, label, bold=True, color=color, fill=fill)
        c = ws.cell(row=row, column=2, value=formula)
        c.font = Font(bold=True, color=color, size=10, name="Arial")
        c.fill = fill
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = '#,##0.00'
        set_cell(row, 4, hint, color="888888", size=9, italic=True, fill=fill)
        ws.row_dimensions[row].height = 20

    def kpi(row, label, formula, fmt="0.00", hint=""):
        set_cell(row, 1, f"  {label}", fill=KPI_BG)
        c = ws.cell(row=row, column=2, value=formula)
        c.font = Font(bold=True, color="4C1D95", size=10, name="Arial")
        c.fill = KPI_BG
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = fmt
        set_cell(row, 4, hint, color="888888", size=9, italic=True, fill=KPI_BG)
        ws.row_dimensions[row].height = 18

    # ── TÍTULO
    ws.merge_cells('A1:D1')
    t = ws.cell(row=1, column=1, value="BALANCE GENERAL  |  VIVA CONSULTING")
    t.fill = BLUE
    t.font = Font(bold=True, color="FFFFFF", size=14, name="Arial")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # ── METADATA
    for i, (lbl, val, hint) in enumerate([
        ("PERIODO","","Ej: Diciembre 2025"),("MES","","Ej: Diciembre"),
        ("AÑO","","Ej: 2025"),("MONEDA","PEN","PEN · USD")], start=2):
        set_cell(i, 1, lbl, bold=True, color="1A3C6E")
        c = ws.cell(row=i, column=2, value=val)
        c.font = Font(color="0000FF", size=10, name="Arial")
        c.alignment = Alignment(horizontal="left", vertical="center")
        set_cell(i, 4, hint, color="AAAAAA", size=9, italic=True)
        ws.row_dimensions[i].height = 18

    ws.row_dimensions[6].height = 6

    # ── CABECERAS
    for col, lbl in [(1,"CONCEPTO"),(2,"MONTO (S/)"),(4,"NOTAS / REFERENCIA")]:
        c = ws.cell(row=7, column=col, value=lbl)
        c.fill = BLUE
        c.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        c.alignment = Alignment(horizontal="center" if col==2 else "left", vertical="center")
    ws.row_dimensions[7].height = 22

    # ════ ACTIVO ════════════════════════════════════════════
    sec_hdr(8,  "▸  ACTIVO CORRIENTE")
    inp(9,  "CAJA Y BANCOS",              "Efectivo, cuentas corrientes, ahorros")
    inp(10, "CUENTAS POR COBRAR",         "Clientes y otras cuentas por cobrar")
    inp(11, "INVENTARIOS",                "Mercadería, materias primas, productos")
    inp(12, "OTROS ACTIVOS CORRIENTES",   "Prepagos, anticipos, crédito tributario")
    frm(13, "TOTAL ACTIVO CORRIENTE",     "=SUM(B9:B12)",    "Suma Activo Corriente", is_main=True)

    sec_hdr(14, "▸  ACTIVO NO CORRIENTE")
    inp(15, "INMUEBLE, MAQUINARIA Y EQUIPO", "Valor histórico / costo de adquisición")
    inp(16, "DEPRECIACIÓN ACUMULADA",        "Ingresa como NEGATIVO", negative=True)
    inp(17, "ACTIVOS INTANGIBLES",           "Software, licencias, marcas, goodwill")
    inp(18, "OTROS ACTIVOS NO CORRIENTES",   "Inversiones LP, otros activos")
    frm(19, "TOTAL ACTIVO NO CORRIENTE",  "=SUM(B15:B18)",   "Suma Activo No Corriente", is_main=True)
    frm(20, "TOTAL ACTIVO",               "=B13+B19",        "Activo Corriente + No Corriente", is_main=True)

    ws.row_dimensions[21].height = 8

    # ════ PASIVO ════════════════════════════════════════════
    sec_hdr(22, "▸  PASIVO CORRIENTE",    fill=RED)
    inp(23, "CUENTAS POR PAGAR",          "Proveedores y otros acreedores CP")
    inp(24, "PRÉSTAMOS CORTO PLAZO",      "Deuda financiera vencimiento < 1 año")
    inp(25, "OTROS PASIVOS CORRIENTES",   "Tributos, remuneraciones, adelantos")
    frm(26, "TOTAL PASIVO CORRIENTE",     "=SUM(B23:B25)",   "Suma Pasivo Corriente", is_main=True)

    sec_hdr(27, "▸  PASIVO NO CORRIENTE", fill=RED)
    inp(28, "DEUDA LARGO PLAZO",          "Préstamos bancarios vencimiento > 1 año")
    inp(29, "OTROS PASIVOS NO CORRIENTES","Provisiones LP, otros pasivos")
    frm(30, "TOTAL PASIVO NO CORRIENTE",  "=SUM(B28:B29)",   "Suma Pasivo No Corriente", is_main=True)
    frm(31, "TOTAL PASIVO",               "=B26+B30",        "Pasivo Corriente + No Corriente", is_main=True)

    ws.row_dimensions[32].height = 8

    # ════ PATRIMONIO ════════════════════════════════════════
    sec_hdr(33, "▸  PATRIMONIO NETO",     fill=GREEN)
    inp(34, "CAPITAL SOCIAL",             "Capital aportado por los accionistas")
    inp(35, "RESERVAS",                   "Reserva legal y otras reservas")
    inp(36, "UTILIDADES RETENIDAS",       "Resultados acumulados de períodos anteriores")
    inp(37, "RESULTADO DEL EJERCICIO",    "Utilidad o pérdida del período actual")
    frm(38, "TOTAL PATRIMONIO",           "=SUM(B34:B37)",   "Suma Patrimonio Neto", is_main=True)
    frm(39, "TOTAL PASIVO + PATRIMONIO",  "=B31+B38",        "Debe igualar TOTAL ACTIVO", is_main=True)

    ws.row_dimensions[40].height = 8

    # ════ VERIFICACIÓN DE CUADRE ════════════════════════════
    ws.merge_cells('A41:D41')
    chk = ws.cell(row=41, column=1,
        value='=IF(ABS(B20-B39)<0.01,"✅  BALANCE CUADRA  —  Activo = Pasivo + Patrimonio  ✅","❌  NO CUADRA  |  Diferencia: "&TEXT(B20-B39,"#,##0.00")&"  — Revisa los datos")')
    chk.font = Font(bold=True, color="065F46", size=11, name="Arial")
    chk.fill = CHK_OK
    chk.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[41].height = 26

    ws.row_dimensions[42].height = 10

    # ════ KPIs FINANCIEROS ══════════════════════════════════
    sec_hdr(43, "▸  KPIs FINANCIEROS  (calculados automáticamente)", fill=PURPLE)
    for col, lbl in [(1,"INDICADOR"),(2,"VALOR"),(4,"BENCHMARK / INTERPRETACIÓN")]:
        c = ws.cell(row=44, column=col, value=lbl)
        c.fill = KPI_BG
        c.font = Font(bold=True, color="4C1D95", size=9, name="Arial")
        c.alignment = Alignment(horizontal="center" if col==2 else "left", vertical="center")
    ws.row_dimensions[44].height = 18

    kpi(45, "Razón Corriente",          '=IF(B26>0,B13/B26,0)',              "0.00x", "AC / PC  ·  ≥ 1.5 saludable  |  < 1.0 riesgo de liquidez")
    kpi(46, "Liquidez Ácida",           '=IF(B26>0,(B13-B11)/B26,0)',        "0.00x", "(AC - Inventarios) / PC  ·  ≥ 1.0 óptimo")
    kpi(47, "Liquidez Absoluta",        '=IF(B26>0,B9/B26,0)',               "0.00x", "Caja / PC  ·  ≥ 0.20 aceptable")
    kpi(48, "Capital de Trabajo Neto",  "=B13-B26",                          '#,##0.00', "AC - PC  ·  Positivo = capacidad operativa")
    kpi(49, "Endeudamiento Total",      '=IF(B20>0,B31/B20,0)',              "0.0%",  "Pasivo / Activo  ·  < 60% saludable")
    kpi(50, "Deuda / Patrimonio",       '=IF(B38>0,B31/B38,0)',              "0.00x", "Apalancamiento  ·  ≤ 1.0 conservador")
    kpi(51, "Deuda LP / Activo Total",  '=IF(B20>0,B30/B20,0)',              "0.0%",  "Pasivo LP / Activo  ·  estructura financiera")
    kpi(52, "Activo Corriente %",       '=IF(B20>0,B13/B20,0)',              "0.0%",  "Proporción activos líquidos")
    kpi(53, "Patrimonio / Activo",      '=IF(B20>0,B38/B20,0)',              "0.0%",  "Solvencia  ·  > 40% saludable")
    kpi(54, "Multiplicador Patrimonio", '=IF(B38>0,B20/B38,0)',              "0.00x", "Activo / Patrimonio  ·  Apalancamiento implícito")

    ws.row_dimensions[55].height = 10

    # ── LEYENDA
    ws.merge_cells('A56:D56')
    leg = ws.cell(row=56, column=1,
        value="LEYENDA:  🔵 Celdas AZULES = ingresa tus datos   ·   ⚫ Celdas OSCURAS = calculadas por fórmula   ·   ✅ Verificación automática de cuadre   ·   🟣 KPIs = solo lectura")
    leg.font = Font(name="Arial", size=9, italic=True, color="666666")
    leg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[56].height = 22

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ─────────────────────────────────────────────────────────
# API: ESTADO DE RESULTADOS
# ─────────────────────────────────────────────────────────

@app.route("/api/estados-resultados/template")
@login_required
def api_er_template():
    out = _generate_er_template()
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="PLANTILLA_ESTADO_RESULTADOS_VIVA.xlsx")


@app.route("/api/estados-resultados", methods=["GET"])
@login_required
def api_er_list():
    conn = get_connection()
    rows = rows_to_list(conn.execute(
        "SELECT * FROM estados_resultados WHERE cliente_id=? ORDER BY anio DESC, created_at DESC",
        (cid(),)
    ).fetchall())
    conn.close()
    return jsonify(rows)


@app.route("/api/estados-resultados/<int:rid>", methods=["GET"])
@login_required
def api_er_get(rid):
    conn = get_connection()
    row = row_to_dict(conn.execute(
        "SELECT * FROM estados_resultados WHERE id=? AND cliente_id=?", (rid, cid())
    ).fetchone())
    conn.close()
    return jsonify(row or {})


@app.route("/api/estados-resultados/<int:rid>", methods=["DELETE"])
@login_required
def api_er_delete(rid):
    conn = get_connection()
    conn.execute("DELETE FROM estados_resultados WHERE id=? AND cliente_id=?", (rid, cid()))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/estados-resultados/importar", methods=["POST"])
@login_required
def api_er_importar():
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    file = request.files["file"]
    if not allowed_file(file.filename, ALLOWED_EXCEL):
        return jsonify({"error": "Solo se aceptan .xlsx, .xls, .csv"}), 400

    filename = f"{uuid.uuid4()}_{file.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    parsed = _parse_financial_excel(filepath, _ER_MAP)
    try:
        os.remove(filepath)
    except OSError:
        pass
    if "error" in parsed:
        return jsonify({"error": parsed["error"]}), 500

    if not parsed.get("periodo_label"):
        parsed["periodo_label"] = request.form.get("periodo", "Sin período")
    if not parsed.get("anio"):
        parsed["anio"] = datetime.now().year

    # Auto-compute totals if missing
    p = parsed
    if not p.get("total_ingresos"):
        p["total_ingresos"] = p.get("ventas_netas", 0) + p.get("otros_ingresos", 0)
    if not p.get("utilidad_bruta"):
        p["utilidad_bruta"] = p.get("total_ingresos", 0) - p.get("costo_ventas", 0)
    if not p.get("total_gastos_operativos"):
        p["total_gastos_operativos"] = p.get("gastos_administrativos", 0) + p.get("gastos_ventas", 0)
    if not p.get("ebitda"):
        p["ebitda"] = p.get("utilidad_bruta", 0) - p.get("total_gastos_operativos", 0)
    if not p.get("ebit"):
        p["ebit"] = p.get("ebitda", 0) - p.get("depreciacion_amortizacion", 0)
    if not p.get("utilidad_antes_impuestos"):
        p["utilidad_antes_impuestos"] = p.get("ebit", 0) - p.get("gastos_financieros", 0) - p.get("otros_gastos_netos", 0)
    if not p.get("utilidad_neta"):
        p["utilidad_neta"] = p.get("utilidad_antes_impuestos", 0) - p.get("impuesto_renta", 0)

    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        INSERT INTO estados_resultados
        (cliente_id, periodo_label, mes, anio, moneda,
         ventas_netas, otros_ingresos, total_ingresos,
         costo_ventas, utilidad_bruta,
         gastos_administrativos, gastos_ventas, total_gastos_operativos,
         ebitda, depreciacion_amortizacion, ebit,
         gastos_financieros, otros_gastos_netos,
         utilidad_antes_impuestos, impuesto_renta, utilidad_neta,
         archivo_origen)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        cid(),
        p.get("periodo_label"), p.get("mes"), p.get("anio"), p.get("moneda", "PEN"),
        p.get("ventas_netas", 0), p.get("otros_ingresos", 0), p.get("total_ingresos", 0),
        p.get("costo_ventas", 0), p.get("utilidad_bruta", 0),
        p.get("gastos_administrativos", 0), p.get("gastos_ventas", 0), p.get("total_gastos_operativos", 0),
        p.get("ebitda", 0), p.get("depreciacion_amortizacion", 0), p.get("ebit", 0),
        p.get("gastos_financieros", 0), p.get("otros_gastos_netos", 0),
        p.get("utilidad_antes_impuestos", 0), p.get("impuesto_renta", 0), p.get("utilidad_neta", 0),
        file.filename,
    ))
    new_id = c.lastrowid
    conn.commit()
    conn.close()

    return jsonify({"success": True, "id": new_id, "periodo_label": p.get("periodo_label"), "data": p})


@app.route("/api/estados-resultados/exportar")
@login_required
def api_er_exportar():
    import pandas as pd
    import io as _io
    conn = get_connection()
    rows = rows_to_list(conn.execute(
        "SELECT * FROM estados_resultados WHERE cliente_id=? ORDER BY anio, created_at",
        (cid(),)
    ).fetchall())
    conn.close()
    if not rows:
        return jsonify({"error": "Sin datos"}), 404
    df = pd.DataFrame(rows)
    out = _io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Estado de Resultados", index=False)
    out.seek(0)
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"VIVA_EstadoResultados_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ─────────────────────────────────────────────────────────
# API: BALANCE GENERAL
# ─────────────────────────────────────────────────────────

@app.route("/api/balance-general/template")
@login_required
def api_bg_template():
    out = _generate_bg_template()
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="PLANTILLA_BALANCE_GENERAL_VIVA.xlsx")


@app.route("/api/balance-general", methods=["GET"])
@login_required
def api_bg_list():
    conn = get_connection()
    rows = rows_to_list(conn.execute(
        "SELECT * FROM balance_general WHERE cliente_id=? ORDER BY anio DESC, created_at DESC",
        (cid(),)
    ).fetchall())
    conn.close()
    return jsonify(rows)


@app.route("/api/balance-general/<int:rid>", methods=["GET"])
@login_required
def api_bg_get(rid):
    conn = get_connection()
    row = row_to_dict(conn.execute(
        "SELECT * FROM balance_general WHERE id=? AND cliente_id=?", (rid, cid())
    ).fetchone())
    conn.close()
    return jsonify(row or {})


@app.route("/api/balance-general/<int:rid>", methods=["DELETE"])
@login_required
def api_bg_delete(rid):
    conn = get_connection()
    conn.execute("DELETE FROM balance_general WHERE id=? AND cliente_id=?", (rid, cid()))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/balance-general/importar", methods=["POST"])
@login_required
def api_bg_importar():
    if "file" not in request.files:
        return jsonify({"error": "No se envió archivo"}), 400
    file = request.files["file"]
    if not allowed_file(file.filename, ALLOWED_EXCEL):
        return jsonify({"error": "Solo se aceptan .xlsx, .xls, .csv"}), 400

    filename = f"{uuid.uuid4()}_{file.filename}"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    parsed = _parse_financial_excel(filepath, _BG_MAP)
    try:
        os.remove(filepath)
    except OSError:
        pass
    if "error" in parsed:
        return jsonify({"error": parsed["error"]}), 500

    if not parsed.get("periodo_label"):
        parsed["periodo_label"] = request.form.get("periodo", "Sin período")
    if not parsed.get("anio"):
        parsed["anio"] = datetime.now().year

    p = parsed
    # Auto-compute totals if missing
    if not p.get("total_activo_corriente"):
        p["total_activo_corriente"] = sum(p.get(k, 0) for k in
            ["caja_bancos", "cuentas_cobrar", "inventarios", "otros_ac"])
    if not p.get("total_activo_no_corriente"):
        p["total_activo_no_corriente"] = sum(p.get(k, 0) for k in
            ["inmueble_maquinaria", "depreciacion_acumulada", "activos_intangibles", "otros_anc"])
    if not p.get("total_activo"):
        p["total_activo"] = p.get("total_activo_corriente", 0) + p.get("total_activo_no_corriente", 0)
    if not p.get("total_pasivo_corriente"):
        p["total_pasivo_corriente"] = sum(p.get(k, 0) for k in
            ["cuentas_pagar", "prestamos_cp", "otros_pc"])
    if not p.get("total_pasivo_no_corriente"):
        p["total_pasivo_no_corriente"] = sum(p.get(k, 0) for k in ["deuda_lp", "otros_pnc"])
    if not p.get("total_pasivo"):
        p["total_pasivo"] = p.get("total_pasivo_corriente", 0) + p.get("total_pasivo_no_corriente", 0)
    if not p.get("total_patrimonio"):
        p["total_patrimonio"] = sum(p.get(k, 0) for k in
            ["capital_social", "reservas", "utilidades_retenidas", "resultado_ejercicio"])
    if not p.get("total_pasivo_patrimonio"):
        p["total_pasivo_patrimonio"] = p.get("total_pasivo", 0) + p.get("total_patrimonio", 0)

    conn = get_connection()
    c = conn.cursor()
    c.execute("""
        INSERT INTO balance_general
        (cliente_id, periodo_label, mes, anio, moneda,
         caja_bancos, cuentas_cobrar, inventarios, otros_ac, total_activo_corriente,
         inmueble_maquinaria, depreciacion_acumulada, activos_intangibles, otros_anc,
         total_activo_no_corriente, total_activo,
         cuentas_pagar, prestamos_cp, otros_pc, total_pasivo_corriente,
         deuda_lp, otros_pnc, total_pasivo_no_corriente, total_pasivo,
         capital_social, reservas, utilidades_retenidas, resultado_ejercicio,
         total_patrimonio, total_pasivo_patrimonio, archivo_origen)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        cid(),
        p.get("periodo_label"), p.get("mes"), p.get("anio"), p.get("moneda", "PEN"),
        p.get("caja_bancos", 0), p.get("cuentas_cobrar", 0), p.get("inventarios", 0),
        p.get("otros_ac", 0), p.get("total_activo_corriente", 0),
        p.get("inmueble_maquinaria", 0), p.get("depreciacion_acumulada", 0),
        p.get("activos_intangibles", 0), p.get("otros_anc", 0),
        p.get("total_activo_no_corriente", 0), p.get("total_activo", 0),
        p.get("cuentas_pagar", 0), p.get("prestamos_cp", 0), p.get("otros_pc", 0),
        p.get("total_pasivo_corriente", 0),
        p.get("deuda_lp", 0), p.get("otros_pnc", 0),
        p.get("total_pasivo_no_corriente", 0), p.get("total_pasivo", 0),
        p.get("capital_social", 0), p.get("reservas", 0),
        p.get("utilidades_retenidas", 0), p.get("resultado_ejercicio", 0),
        p.get("total_patrimonio", 0), p.get("total_pasivo_patrimonio", 0),
        file.filename,
    ))
    new_id = c.lastrowid
    conn.commit()
    conn.close()
    return jsonify({"success": True, "id": new_id, "periodo_label": p.get("periodo_label"), "data": p})


@app.route("/api/balance-general/exportar")
@login_required
def api_bg_exportar():
    import pandas as pd
    import io as _io
    conn = get_connection()
    rows = rows_to_list(conn.execute(
        "SELECT * FROM balance_general WHERE cliente_id=? ORDER BY anio, created_at",
        (cid(),)
    ).fetchall())
    conn.close()
    if not rows:
        return jsonify({"error": "Sin datos"}), 404
    df = pd.DataFrame(rows)
    out = _io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="Balance General", index=False)
    out.seek(0)
    return send_file(out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"VIVA_BalanceGeneral_{datetime.now().strftime('%Y%m%d')}.xlsx")


@app.errorhandler(404)
def not_found(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": "Recurso no encontrado"}), 404
    return render_template("login.html", error=None), 404


@app.errorhandler(500)
def server_error(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": "Error interno del servidor"}), 500
    return render_template("login.html", error="Error interno. Por favor recarga la página."), 500


@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "Archivo demasiado grande. Máximo 50 MB."}), 413


if __name__ == "__main__":
    app.run(debug=True, port=5050, host="0.0.0.0")
